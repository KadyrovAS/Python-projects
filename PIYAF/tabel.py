import openpyxl as op
import excel_format

fl = open("tabel.ini", "r", encoding="utf-8")
lines = fl.readlines()
path = None

for item in lines:
    com = item.split("=")
    if com[0].strip() == "path":
        path = com[1].strip()

if path is None:
    print("Отсутствует значение 'папка' в ini файле")


pathShablon = path + "\\" + "Шаблон.xlsx"
try:
    excel_format.create_holidays(pathShablon)
except Exception:
    print('Закройте файл "Шаблон.xlsx"')
    exit()

excelShablon = op.open(filename = pathShablon, data_only = False)
# sheetnames = excelShablon.sheetnames
sheet = excelShablon['Работники']

spisok = []
nColumn = 1
while sheet.cell(row = 1, column = nColumn).value is not None:
    nColumn += 1

nColumn -= 1
row = 2
while sheet.cell(row = row, column = 1).value is not None:
    person = [sheet.cell(row = row, column=col).value for col in range(1, nColumn + 1)]
    if person[3] is None:
        person[3] = " "
    spisok.append(person)
    row += 1

excel_format.sortSpisok(spisok, True)

row = 2
for person in spisok:
    col = 1
    for val in person:
        sheet.cell(row = row, column = col).value = val
        col += 1
    row += 1

excel_format.sortSpisok(spisok)

excel_format.border(sheet, True)
excelShablon.save(pathShablon)
excelShablon.close

excel_format.tabelCreate(spisok, path)
