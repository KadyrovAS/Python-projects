import datetime
import openpyxl as op
import openpyxl.styles as st
import os.path
import createTabel as create

holidays = []
def to_box_border(sheet, row1, row2, col1, col2):
    thin = st.Side(border_style="double", color="000000")  # параметры тонкой линии
    for col in range(col1, col2 + 1):
        sheet.cell(row=row1, column=col).border = st.Border(top=thin)
        sheet.cell(row=row2, column=col).border = st.Border(bottom=thin)
    for row in range(row1, row2 + 1):
        sheet.cell(row=row, column=col1).border = st.Border(left=thin)
        sheet.cell(row=row, column=col2).border = st.Border(right=thin)
    sheet.cell(row=row1, column=col1).border = st.Border(left=thin, top=thin)
    sheet.cell(row=row1, column=col2).border = st.Border(right=thin, top=thin)
    sheet.cell(row=row2, column=col1).border = st.Border(left=thin, bottom=thin)
    sheet.cell(row=row2, column=col2).border = st.Border(right=thin, bottom=thin)


def to_some_borders(sheet, row, column1, column2):
    thin = st.Side(border_style="thin", color="000000")  # параметры тонкой линии
    for col in range(column1, column2 + 1):
        sheet.cell(row=row, column=col).border = st.Border(bottom=thin)
def border(sheet: op.worksheet, arg = False, nRowStart = 0, nRowFinis = 0, nColStart = 0, nColFinis = 0):
    # arg = True - бордюр рисуется
    #arg = False - бордюр убирается
    #Если nRowStart, nRowFinis, nColStart, nColFinis не указаны, то применяется ко всему листу
    #Если nRowStart, nRowFinis, nColStart, nColFinis указаны, то применяется только к указанному диапазону
    #Диапазон листа рассчитывается следующим образом:
    # - количество столбцов считается в 1-й строке до первой пустой клетки
    # - количество строк считается по 1-му столбцу до первой пустой клетки
    thin = st.Side(border_style="thin", color="000000") #параметры тонкой линии
    if nRowStart == 0:
        #Определяется диапазон листа
        nRowStart = 1
        nRowFinis = 1
        nColStart = 1
        nColFinis = 1
        #считаем количество строк по 1-му столбцу
        while sheet.cell(row = nRowFinis, column = 1).value is not None:
            nRowFinis += 1
        #считаем количество столбцов по 1-й строке
        while sheet.cell(row = 1, column = nColFinis).value is not None:
            nColFinis += 1
        nRowFinis -= 1
        nColFinis -= 1

    #применяем правило к диапазону
    for row in range(nRowStart, nRowFinis + 1):
        for col in range(nColStart, nColFinis + 1):
            currentCell = sheet.cell(row = row, column = col)
            if arg:
                currentCell.border = st.Border(top=thin, left=thin, right=thin, bottom=thin)
            else:
                currentCell.border=st.Border(top=None, left=None, right=None, bottom=None)

def getMonth(month = 0, arg = 'Name', year = 0):
    if year == 0:
        year = datetime.datetime.now().year

    if month == 0:
        month = datetime.datetime.now().month

    months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
              'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    monthsR = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня',
               'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября', 'Декабря']
    days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    if year % 4 == 0:
        days[1] = 29

    if arg == "Name":
        return months[month - 1]
    elif arg == "NameR":
        return monthsR[month - 1]
    elif arg == "count":
        return days[month - 1]
    elif arg == "month":
        return month
    elif arg == "date":
        dt = datetime.datetime.now()
        return f'{dt.day}:{dt.month}:{dt.year}'
    elif arg == "year":
        return year


def tabelCreate(spisok, path: str, year = 0, month = 0):
    if year == 0:
        year = datetime.datetime.now().year
        month = datetime.datetime.now().month
    try:
        filename = f"{path}\\Табель {year}.xlsx"
    except Exception:
        print(f'Закройте файл "{filename}"')
        return

    if os.path.exists(filename):
        #Открываем существующий файл
        wbFileTabel = op.open(filename=filename, data_only=False)
    else:
        #Создаем файл
        wbFileTabel = op.Workbook()

    sheets = wbFileTabel.sheetnames
    currentSheet = None
    sheetName = f"{year} {getMonth(month)}"

    for sheet in sheets:
        if sheet == 'Sheet':
            currentSheet = sheet
            wbFileTabel[currentSheet].title = sheetName
            create.createNewTable(spisok, wbFileTabel[sheetName], path)
        elif sheet == sheetName:
            currentSheet = sheet
            break

    if currentSheet is None:
        wbFileTabel.create_sheet(sheetName)
        create.createNewTable(spisok, wbFileTabel[sheetName], path)
    else:
        print(f'Файл {filename} с табелем существует!')
        print("Изменения в существующий файл не вносились!")

    try:
        wbFileTabel.save(filename=filename)
    except Exception:
        print(f'Закройте файл "{filename}"')
        return

def sortSpisok(spisok : list, brench = False):
    if brench:
        start = 0
    else:
        start = 1
    #Сортировка списка по алфавиту
    for i in range(len(spisok) - 1):
        for k in range(i + 1, len(spisok)):
            for j in range(start, 4):
                if spisok[i][j] > spisok[k][j]:
                    spisok[i], spisok[k] = spisok[k], spisok[i]
                    break
                elif spisok[i][j] == spisok[k][j]:
                    continue
                else:
                    break

def create_holidays(pathShablon):
    global holidays
    excelShablon = op.open(filename=pathShablon, data_only=False)
    sheet = excelShablon['Выходные']
    holidays = []
    row = 1
    while sheet.cell(row = row, column = 1).value is not None:
        lines = sheet.cell(row = row, column = 1).value.split()
        if lines[0] == "Выходные":
            row += 1
            continue
        sheet.cell(row = row, column = 1).value = ""
        holidays.append(lines[0])
        row += 1

    month = getMonth(arg = 'month') - 1
    if month == 0:
        month = 12

    for m in range(len(holidays) - 1, -1, -1):
        dmy = holidays[m].split(":")
        mm = int(dmy[1])
        if mm < month or mm == 12 and month != 12:
            holidays.pop(m)

    for m in range(month, month + 2):
        if m == month == 12:
            year = getMonth(arg='year') - 1
        else:
            year = getMonth(arg='year')
        n = getMonth(arg='count', month = m, year=year)
        found = False
        for holiday in holidays:
            dmy = holiday.split(":")
            if int(dmy[1]) == m:
                found = True
                break
        if found:
            continue
        for day in range(1, n + 1):
            date = datetime.datetime(year, m, day)
            if date.weekday() + 1 == 6 or date.weekday() + 1 == 7:
                dd = date.day
                mm = date.month
                yy = date.year
                dd = f'0{dd}' if dd < 10 else f'{dd}'
                mm = f'0{mm}' if mm < 10 else f'{mm}'
                dt = f'{dd}:{mm}:{yy}'

                if not found:
                    holidays.append(dt)

    for i in range(len(holidays) - 1):
        dmy = holidays[i].split(":")
        dt1 = dmy[2] + dmy[1] + dmy[0]
        for k in range(i + 1, len(holidays)):
            dmy = holidays[k].split(":")
            dt2 = dmy[2] + dmy[1] + dmy[0]
            if dt1 > dt2:
                holidays[i], holidays[k] = holidays[k], holidays[i]
    row = 1
    sheet.cell(row=row, column=1).value = f"Выходные {getMonth(month=month, arg='NameR')}"
    for dt in holidays:
        dts = dt.split(":")
        mm = int(dts[1])
        row += 1
        if mm != month:
            month = mm
            sheet.cell(row=row, column=1).value = f"Выходные {getMonth(month=month, arg='NameR')}"
            row += 1
        sheet.cell(row = row, column = 1).value = dt
    excelShablon.save(pathShablon)

def is_holiday(dt=None, day = 0):
    global holidays
    if dt == None:
        dt = f'{day}:{getMonth(arg="month")}:{getMonth(arg="year")}'

    dt_list = list(map(int, dt.split(':')))
    dd = f'0{dt_list[0]}' if dt_list[0] < 10 else dt_list[0]
    mm = f'0{dt_list[1]}' if dt_list[1] < 10 else dt_list[1]

    dt = f'{dd}:{mm}:{getMonth(arg="year")}'
    return dt in holidays

def get_pattern_fill(dt = None, day = 0):
    color = {
        "Holiday": st.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type = "solid"),  # Dark Green
        "Workday": st.PatternFill(start_color=None, end_color=None, fill_type = "none") # Red
    }
    if dt == None:
        dt = f'{day}:{getMonth(arg="month")}:{getMonth(arg="year")}'
    if is_holiday(dt):
        return color['Holiday']
    else:
        return color['Workday']
