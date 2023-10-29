import openpyxl as op
import openpyxl.styles as st
from openpyxl.utils import get_column_letter

import excel_format
def createHead(workSheet, rowStart, n):
    # Объединение ячеек
    workSheet.merge_cells(f'B{rowStart}:C{rowStart}')

    letter = get_column_letter(n + 6)
    workSheet.merge_cells(f'E{rowStart}:{letter}{rowStart}')

    workSheet.merge_cells(f'A{rowStart}:A{rowStart + 1}')
    workSheet.merge_cells(f'D{rowStart}:D{rowStart + 1}')

    # Заполняем шапку таблицы
    workSheet.cell(row=rowStart, column=1).value = "Фамилия, имя, отчество"
    workSheet.cell(row=rowStart, column=1).alignment = st.Alignment(wrap_text=True,
                                                                    horizontal='center', vertical='center')
    workSheet.cell(row=rowStart, column=2).value = "Учетный номер"
    workSheet.cell(row=rowStart + 1, column=2).value = "Табельный номер"
    workSheet.cell(row=rowStart + 1, column=3).value = "Премия"
    workSheet.cell(row=rowStart, column=4).value = "Должность (профессия)"
    workSheet.cell(row=rowStart, column=4).alignment = st.Alignment(wrapText=True,
                                                                    horizontal='center', vertical='center')
    workSheet.cell(row=rowStart, column=5).value = "Числа месяца"
    col = 4
    for i in range(n):
        col += 1
        if i == 15:
            workSheet.cell(row=rowStart + 1, column=col).value = "Итого дней (часов) явок (неявок) с 1 по 15"
            workSheet.cell(row=rowStart + 1, column=col).alignment = st.Alignment(wrap_text=True,
                                                                                  horizontal='center',
                                                                                  vertical='center')
            col += 1
        workSheet.cell(row=rowStart + 1, column=col).value = i + 1
        workSheet.cell(row=rowStart + 1, column=col).fill = excel_format.get_pattern_fill(day=i+1)
        workSheet.cell(row=rowStart + 2, column=col).fill = excel_format.get_pattern_fill(day=i+1)


    col += 1
    workSheet.cell(row=rowStart + 1, column=col).value = "Всего дней (часов) явок (неявок) за месяц"
    workSheet.cell(row=rowStart + 1, column=col).alignment = st.Alignment(wrap_text=True,
                                                                          horizontal='center', vertical='center')

    workSheet[f'B{rowStart + 1}'].alignment = st.Alignment(textRotation=90, horizontal='center', vertical='center')
    workSheet[f'C{rowStart + 1}'].alignment = st.Alignment(textRotation=90, horizontal='center', vertical='center')
    alignment = st.Alignment(horizontal='center', vertical='center')

    for i in range(1, col + 1):
        workSheet.cell(row=rowStart + 2, column=i).value = i
        workSheet.cell(row=rowStart + 2, column=i).alignment = alignment


def createNewTable(spisok, workSheet, path):
    font = st.Font(
        size=8,
        bold=False,
        italic=False,
        color="000000",
        name="Arial"
    )
    alignment = st.Alignment(wrap_text=True, horizontal='center', vertical='center')

    # ориентация страницы
    workSheet.page_setup.orientation = workSheet.ORIENTATION_LANDSCAPE
    workSheet.page_setup.paperSize = workSheet.PAPERSIZE_A4
    # зумирование по ширине
    workSheet.sheet_properties.pageSetUpPr.fitToPage = True
    workSheet.page_setup.fitToWidth = False

    for row in range(1, 1000):
        for col in range(1, 40):
            workSheet.cell(row=row, column=col).font = font
            workSheet.cell(row=row, column=col).alignment = alignment

    pathShablon = f"{path}\\Шаблон.xlsx"
    wbShablon = op.open(filename=pathShablon, data_only=False)
    ws = wbShablon['Настройки']
    row = 2
    spr = dict()
    while ws.cell(row=row, column=1).value is not None:
        cl = ws.cell(row=row, column=1).value
        clValue = ws.cell(row=row, column=2).value
        clValue = float(clValue)
        if cl == "ФИО":
            spr.update({"FIO": clValue})
        elif cl == "Табельный номер":
            spr.update({'Tabel': clValue})
        elif cl == "Премия":
            spr.update({'Prem': clValue})
        elif cl == "Должность":
            spr.update({'Dol': clValue})
        elif cl == "Дата":
            spr.update({'Data': clValue})
        elif cl == "Итого":
            spr.update({'Itog': clValue})
        elif cl == "Количество":
            spr.update({'Count': clValue})
        row += 1

    workSheet.column_dimensions['A'].width = spr['FIO']  # фамилия имя отчество
    workSheet.column_dimensions['B'].width = spr['Tabel']  # Табельный номер
    workSheet.column_dimensions['C'].width = spr['Prem']  # Премия
    workSheet.column_dimensions['D'].width = spr['Dol']  # Должность
    col = 4
    for i in range(15):
        col += 1
        letter = get_column_letter(col)
        workSheet.column_dimensions[letter].width = spr['Data']  # Даты

    col += 1
    letter = get_column_letter(col)
    workSheet.column_dimensions[letter].width = spr['Itog']

    n = excel_format.getMonth(arg='count')
    for i in range(16, n + 1):
        col += 1
        letter = get_column_letter(col)
        workSheet.column_dimensions[letter].width = spr['Data']  # Даты

    col += 1
    letter = get_column_letter(col)
    workSheet.column_dimensions[letter].width = spr['Itog']

    createHead(workSheet, 11, n)

    # Заполняем таблицу списком
    font1 = st.Font(
        size=8,
        bold=True,
        italic=False,
        color="000000",
        name="Arial"
    )
    row = 14
    for i in range(len(spisok)):
        worker = spisok[i]
        col = 5
        for k in range(1, n + 1):
            if k == 16:
                workSheet.cell(row=row, column=col).value = f"=COUNTA(E{row}:S{row})"
                workSheet.cell(row=row+1, column=col).value = f"=SUM(E{row}:S{row})"
                workSheet.cell(row=row, column=col).font=font1
                workSheet.cell(row=row+1, column=col).font=font1

                col += 1
            workSheet.cell(row=row, column=col).fill = excel_format.get_pattern_fill(day=k)
            workSheet.cell(row=row+1, column=col).fill = excel_format.get_pattern_fill(day=k)
            if excel_format.is_holiday(day=k):
                workSheet.cell(row=row+1, column=col).value='В'
            else:
                workSheet.cell(row=row, column=col).value = worker[6]
                workSheet.cell(row=row + 1, column=col).value = worker[7]

            col += 1

        column0_letter = get_column_letter(20)
        column1_letter = get_column_letter(21)
        column2_letter = get_column_letter(col - 1)

        workSheet.cell(row=row, column=col).value = f"=COUNTA({column1_letter}{row}:{column2_letter}{row}) +" \
                f"{column0_letter}{row}"
        workSheet.cell(row=row + 1, column=col).value = f"=SUM({column1_letter}{row}:{column2_letter}{row}) +" \
                f"{column0_letter}{row+1}"
        workSheet.cell(row=row, column=col).font = font1
        workSheet.cell(row=row+1, column=col).font = font1

        workSheet.merge_cells(f'A{row}:A{row + 1}')
        workSheet.cell(row=row, column=1).alignment = alignment
        workerName = f'{worker[1]} {worker[2][0]}.' + ('' if worker[3] == '' else (worker[3][0] + '.'))
        workSheet.cell(row=row, column=1).value = workerName

        workSheet.merge_cells(f'B{row}:B{row + 1}')
        workSheet.cell(row=row, column=2).alignment = alignment
        workSheet.cell(row=row, column=2).value = worker[5]

        workSheet.merge_cells(f'D{row}:D{row + 1}')
        workSheet.cell(row=row, column=4).alignment = alignment
        workSheet.cell(row=row, column=4).value = worker[4]

        row += 2
        if (i + 1) % spr['Count'] == 0 and (i + 1) != len(spisok):
            # next_page_horizon, next_page_vertical = workSheet.page_breaks
            # next_page_horizon.append(Break(row))
            createHead(workSheet, row, n)
            row += 3

    excel_format.border(workSheet, True, 11, row - 1, 1, n + 6)
    createMainHead(workSheet)
    createCell(workSheet, row)

def createMainHead(workSheet):
    n = excel_format.getMonth(arg = 'count')
    month = excel_format.getMonth(arg = 'month')
    column = get_column_letter(n + 3)
    alignment = st.Alignment(horizontal='center', vertical='center')
    font1 = st.Font(
        size=14,
        bold=True,
        italic=False,
        color="000000",
        name="Arial"
    )

    workSheet.merge_cells(f'A1:{column}1')
    workSheet.cell(row=1, column=1).value = f'Табель №{month}'
    workSheet.cell(row=1, column=1).alignment = alignment
    workSheet.cell(row=1, column=1).font = font1

    font2 = st.Font(
        size=11,
        bold=True,
        italic=False,
        color="000000",
        name="Arial"
    )
    workSheet.merge_cells(f'A2:{column}2')
    workSheet.cell(row=2, column=1).value = 'учета использования рабочего времени'
    workSheet.cell(row=2, column=1).alignment = alignment
    workSheet.cell(row=2, column=1).font = font2

    column1_letter = get_column_letter(n + 4)
    column2_letter = get_column_letter(n + 6)

    font3 = st.Font(
        size=8,
        bold=False,
        italic=False,
        color="000000",
        name="Arial"
    )

    for row in range(2, 10):
        workSheet.merge_cells(f'{column1_letter}{row}:{column2_letter}{row}')
        workSheet.cell(row = row, column = n + 4).alignment = alignment
        workSheet.cell(row = row, column = n + 4).font = font3

    workSheet.cell(row = 2, column = n + 4).value = "Коды"
    workSheet.cell(row = 3, column = n + 4).value = "0504421"
    workSheet.cell(row = 4, column = n + 4).value = excel_format.getMonth(arg = 'date')
    workSheet.cell(row = 5, column = n + 4).value = "02698654"
    workSheet.cell(row = 8, column = n + 4).value = "0"
    workSheet.cell(row = 9, column = n + 4).value = excel_format.getMonth(arg = 'date')

    column3_letter = get_column_letter(n)
    column4_letter = get_column_letter(n+3)
    column5_letter = 'C'
    column6_letter = get_column_letter(n - 5)
    alignment = st.Alignment(horizontal='right', vertical='center')
    col=n
    for row in range(3, 10):
        workSheet.merge_cells(f'{column3_letter}{row}:{column4_letter}{row}')
        workSheet.cell(row=row, column=col).alignment = alignment
        workSheet.merge_cells(f'{column5_letter}{row}:{column6_letter}{row}')

        if row == 7:
            column3_letter = get_column_letter(n - 4)
            col -= 4

    workSheet.cell(row=3, column=n).value = "Форма по ОКУД"
    workSheet.cell(row=4, column=n).value = "Дата"
    workSheet.cell(row=5, column=n).value = "по ОКПО"
    workSheet.cell(row=8, column=n - 4).value = "Номер корректировки"
    workSheet.cell(row=9, column=n - 4).value = "Дата формирования документа"

    alignment = st.Alignment(horizontal='left', vertical='center')
    workSheet.merge_cells('A5:B5')
    workSheet.merge_cells('A7:B7')
    workSheet.merge_cells('A8:B8')
    workSheet.cell(row=5, column=1).alignment = alignment
    workSheet.cell(row=7, column=1).alignment = alignment
    workSheet.cell(row=8, column=1).alignment = alignment

    workSheet.cell(row=5, column=1).value = "Учреждение"
    workSheet.cell(row=7, column=1).value = "Структурное подразделение"
    workSheet.cell(row=8, column=1).value = "Вид табеля"

    workSheet.cell(row=4, column=3).value = \
        f'за период с 1 по {excel_format.getMonth(arg="count")} {excel_format.getMonth(arg="NameR")} ' + \
        f'{excel_format.getMonth(arg="year")} года'
    excel_format.to_some_borders(workSheet, 4, 11, 17)

    workSheet.cell(row=5, column=3).value = \
        'Федеральное государственное бюджетное учреждение "Петербургский институт ядерной физики им. Б.П. Константинова'
    workSheet.cell(row=5, column=3).alignment = alignment
    excel_format.to_some_borders(workSheet, 5, 3, n - 7)

    workSheet.cell(row=6, column=3).value= \
    'Национального исследовательского центра "Курчатовский институт"'
    workSheet.cell(row=6, column=3).alignment = alignment
    excel_format.to_some_borders(workSheet, 6, 3, 14)

    workSheet.cell(row=7, column=3).value = \
    "отдел оптических и информационных технологий отделения перспективных разработок (ООИТ ОПР)"
    workSheet.cell(row=7, column=3).alignment = alignment
    excel_format.to_some_borders(workSheet, 7, 3, 20)

    workSheet.cell(row = 8, column=3).value = "первичный"
    workSheet.cell(row=8, column=3).alignment = alignment
    excel_format.to_some_borders(workSheet, 8, 3, 4)

    workSheet.merge_cells

    excel_format.border(workSheet, True, 2, 9, n + 4, n + 6)

def createCell(workSheet, row):
    font1 = st.Font(
        size=11,
        bold=True,
        italic=False,
        color="000000",
        name="Arial"
    )
    nCol=excel_format.getMonth(arg="count")
    n = nCol // 2
    excel_format.to_box_border(workSheet, row + 1, row + 6, n + 4, nCol + 6)
    column1_letter = get_column_letter(n + 4)
    column2_letter = get_column_letter(nCol + 6)
    workSheet.merge_cells(f'{column1_letter}{row+1}:{column2_letter}{row+1}')
    workSheet.cell(row=row + 1, column=n + 4).value = "Отметка бухгалтерии о принятии настоящего табеля"
    workSheet.cell(row=row + 1, column=n + 4).alignment = st.Alignment(horizontal='center', vertical='center')
    workSheet.cell(row=row + 1, column=n + 4).font = font1

    l = (nCol + 2 - n) // 4
    col = n + 4
    for i in range(4):
        column3_letter = get_column_letter(col)
        if i == 3:
            col4 = nCol + 6
            column4_letter = get_column_letter(col4)
        else:
            col4 = col + l - 1
            column4_letter = get_column_letter(col4)
        workSheet.merge_cells(f'{column3_letter}{row + 3}:{column4_letter}{row + 3}')
        workSheet.merge_cells(f'{column3_letter}{row + 4}:{column4_letter}{row + 4}')
        if i > 0:
            excel_format.to_some_borders(workSheet, row + 3, col, col4 - 1)

        col += l

    workSheet.cell(row=row + 3, column=n + 4).value = "Исполнитель"
    workSheet.cell(row=row + 4, column=n + 4 + l).value = "(должность)"
    workSheet.cell(row=row + 4, column=n + 4 + 2 * l).value = "(подпись)"
    workSheet.cell(row=row + 4, column=n + 4 + 3 * l).value = "(расшифровка подписи)"
    column1_letter = get_column_letter(n + 5)
    column2_letter = get_column_letter(nCol + 5)
    workSheet.merge_cells(f'{column1_letter}{row + 5}:{column2_letter}{row + 5}')
    workSheet.cell(row=row + 5, column=n + 5).value = '"______" __________________________________ 202__'
    workSheet.cell(row=row + 5, column=n + 5).alignment = st.Alignment(horizontal='left', vertical='center')

    font2 = st.Font(
        size=8,
        bold=True,
        italic=False,
        color="000000",
        name="Arial"
    )
    for r in range(row + 1, row + 3):
        workSheet.cell(r, column = 1).alignment = st.Alignment(horizontal='left', vertical='center')
        workSheet.cell(r, column = 1).font = font2
    workSheet.cell(row + 1, column = 1).value = "Ответственный"
    workSheet.cell(row + 2, column = 1).value = "исполнитель"

    for r in range(row + 4, row + 6):
        workSheet.cell(r, column = 1).alignment = st.Alignment(horizontal='left', vertical='center')
        workSheet.cell(r, column = 1).font = font2

    workSheet.cell(row + 4, column = 1).value = "Исполнитель"


    workSheet.merge_cells(f'B{row + 1}:D{row + 1}')
    workSheet.merge_cells(f'B{row + 2}:D{row + 2}')
    excel_format.to_some_borders(workSheet, row + 1, 2, 4)
    workSheet.cell(row = row + 1, column = 2).value = "Заведующий ООИТ"
    workSheet.cell(row = row + 2, column = 2).value = "(должность)"

    workSheet.merge_cells(f'B{row + 4}:D{row + 4}')
    workSheet.merge_cells(f'B{row + 5}:D{row + 5}')
    excel_format.to_some_borders(workSheet, row + 4, 2, 4)
    workSheet.cell(row = row + 4, column = 2).value = "Старший лаборант"
    workSheet.cell(row = row + 5, column = 2).value = "(должность)"

    workSheet.merge_cells(f'F{row + 1}:I{row + 1}')
    workSheet.merge_cells(f'F{row + 2}:I{row + 2}')
    excel_format.to_some_borders(workSheet, row + 1, 6, 9)
    workSheet.cell(row = row + 2, column = 6).value = "(подпись)"

    workSheet.merge_cells(f'F{row + 4}:I{row + 4}')
    workSheet.merge_cells(f'F{row + 5}:I{row + 5}')
    excel_format.to_some_borders(workSheet, row + 4, 6, 9)
    workSheet.cell(row = row + 5, column = 6).value = "(подпись)"

    workSheet.merge_cells(f'K{row + 1}:P{row + 1}')
    workSheet.merge_cells(f'K{row + 2}:P{row + 2}')
    excel_format.to_some_borders(workSheet, row + 1, 11, 16)
    workSheet.cell(row = row + 1, column = 11).value = "А.С. Кадыров"
    workSheet.cell(row = row + 2, column = 11).value = "(расшифровка подписи)"

    workSheet.merge_cells(f'K{row + 4}:P{row + 4}')
    workSheet.merge_cells(f'K{row + 5}:P{row + 5}')
    excel_format.to_some_borders(workSheet, row + 4, 11, 16)
    workSheet.cell(row = row + 4, column = 11).value = "О.Е. Тупиленко"
    workSheet.cell(row = row + 5, column = 11).value = "(расшифровка подписи)"

    n1 = (nCol + 6) //2 - 2
    excel_format.to_box_border(workSheet, row + 8, row + 13, 1, n1)
    n2 = n1 + 2
    excel_format.to_box_border(workSheet, row + 8, row + 13, n2, nCol + 6)

    column1_letter = 'A'
    column2_letter = get_column_letter(n1)
    workSheet.merge_cells(f'{column1_letter}{row + 8}:{column2_letter}{row + 8}')
    workSheet.cell(row = row + 8, column = 1).font = font1
    workSheet.cell(row = row + 8, column = 1).value = \
        "Отметка отдела труда и  заработной платы о проверке настоящего табеля"

    column1_letter = get_column_letter(n2)
    column2_letter = get_column_letter(nCol + 6)
    workSheet.merge_cells(f'{column1_letter}{row + 8}:{column2_letter}{row + 8}')
    workSheet.cell(row = row + 8, column = n2).font = font1
    workSheet.cell(row = row + 8, column = n2).value = "Отметка отдела кадров о проверке настоящего табеля"

    workSheet.cell(row = row + 10, column = 1).value = "Исполнитель"
    workSheet.cell(row = row + 10, column = 1).alignment = st.Alignment(horizontal='left', vertical='center')
    column1_letter = get_column_letter(n2)
    column2_letter = get_column_letter(n2 + 2)
    workSheet.merge_cells(f'{column1_letter}{row + 10}:{column2_letter}{row + 10}')
    workSheet.cell(row = row + 10, column = n2).value = "Исполнитель"
    workSheet.cell(row = row + 10, column = n2).alignment = st.Alignment(horizontal='left', vertical = 'center')

    excel_format.to_some_borders(workSheet, row + 10, 2, 3)
    excel_format.to_some_borders(workSheet, row + 10, n2 + 3, n2 + 5)
    column1_letter = get_column_letter(n2 + 3)
    column2_letter = get_column_letter(n2 + 5)
    workSheet.merge_cells(f'B{row + 11}:C{row + 11}')
    workSheet.cell(row = row + 11, column = 2).value = "(должность)"
    workSheet.merge_cells(f'{column1_letter}{row + 11}:{column2_letter}{row + 11}')
    workSheet.cell(row = row + 11, column = n2 + 3).value = "(должность)"

    column1_letter = get_column_letter(n2 + 7)
    column2_letter = get_column_letter(n2 + 10)
    excel_format.to_some_borders(workSheet, row + 10, 5, 8)
    excel_format.to_some_borders(workSheet, row + 10, n2 + 7, n2 + 10)
    workSheet.merge_cells(f'E{row + 11}:G{row + 11}')
    workSheet.cell(row = row + 11, column = 5).value = "(подпись)"
    workSheet.merge_cells(f'{column1_letter}{row + 11}:{column2_letter}{row + 11}')
    workSheet.cell(row = row + 11, column = n2 + 7).value = "(подпись)"

    column1_letter = get_column_letter(n2 + 12)
    column2_letter = get_column_letter(n2 + 17)
    excel_format.to_some_borders(workSheet, row + 10, 10, 15)
    excel_format.to_some_borders(workSheet, row + 10, n2 + 12, n2 + 17)
    workSheet.merge_cells(f'J{row + 11}:O{row + 11}')
    workSheet.cell(row = row + 11, column = 10).value = "(расшифровка подписи)"
    workSheet.merge_cells(f'{column1_letter}{row + 11}:{column2_letter}{row + 11}')
    workSheet.cell(row = row + 11, column = n2 + 12).value = "(расшифровка подписи)"

    column1_letter = get_column_letter(1)
    column2_letter = get_column_letter(n1)
    workSheet.merge_cells(f'{column1_letter}{row + 12}:{column2_letter}{row + 12}')
    column1_letter = get_column_letter(n1 + 2)
    column2_letter = get_column_letter(nCol + 6)
    workSheet.merge_cells(f'{column1_letter}{row + 12}:{column2_letter}{row + 12}')

    workSheet.cell(row = row + 12, column = 1).value = '"_______"____________________________________ 202___'
    workSheet.cell(row = row + 12, column = n1 + 2).value = '"_______"____________________________________ 202___'
    workSheet.cell(row = row + 12, column = 1).alignment = st.Alignment(horizontal='left', vertical='center')
    workSheet.cell(row = row + 12, column = n1 + 2).alignment = st.Alignment(horizontal='left', vertical='center')
